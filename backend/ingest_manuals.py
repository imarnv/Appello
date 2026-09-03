"""
Technical Manual Ingestion
==========================
Adds the commercial-vehicle technical library to the existing
'kb_ggs_support' Qdrant collection, alongside the original short briefs.

Why this is not ingest_briefs.py
--------------------------------
Three things in this corpus break the original pipeline:

1. It is ADDITIVE. ingest_briefs.py drops the collection on every run; the
   45 original brief vectors have to survive, so we delete per-filename instead.

2. The maintenance schedules are legend-coded tables ("• L L L | L: 24") whose
   symbol definitions live on a different page from the table. Extracting a page
   in isolation produces confident nonsense — a dot column gets read as a
   distance column and every interval shifts one column left. So each document
   gets a LEGEND DISCOVERY pass first, and that legend is injected into the
   prompt for every page of that document.

3. At 2,165 pages a serial run is ~12 hours. Extraction is concurrent and
   rate-limited against the deployment's real 250 RPM / 250k TPM budget.

Vision output is checkpointed to JSONL before anything is embedded, so a failure
during embedding never re-spends the vision budget. Re-running skips pages
already in the checkpoint.

Usage:
    python ingest_manuals.py                # extract (resumable) + embed + upsert
    python ingest_manuals.py --extract-only # stop after vision extraction
    python ingest_manuals.py --embed-only   # embed from an existing checkpoint
    python ingest_manuals.py --files "Owners Manual.pdf,Service Manual.pdf"
    python ingest_manuals.py --tag-legacy   # only re-tag the old eCanter specsheet
"""

from __future__ import annotations

import argparse
import base64
import json
import os
import re
import sys
import threading
import time
import urllib.error
import urllib.request
import uuid
from collections import deque
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import pymupdf
from dotenv import load_dotenv

load_dotenv(Path(__file__).parent / ".env")

# ─── Config ──────────────────────────────────────────────────────────────
SOURCE_DIR = Path(__file__).parent.parent / "E-Canter"
CHECKPOINT = Path(__file__).parent / "ecanter_pages.jsonl"
EXTRACT_DUMP = Path(__file__).parent / "ecanter_extracted_text.txt"

VISION_DEPLOYMENT = os.getenv("ECANTER_VISION_DEPLOYMENT", "gpt-4.1-mini")
AZURE_CHAT_RESOURCE = os.getenv("AZURE_OPENAI_RESOURCE", "").rstrip("/")
VISION_ENDPOINT = (
    f"{AZURE_CHAT_RESOURCE}/openai/deployments/"
    f"{VISION_DEPLOYMENT}/chat/completions?api-version=2025-01-01-preview"
)
AZURE_API_KEY = os.getenv("AZURE_OPENAI_CHAT_API_KEY", "")

AZURE_EMBEDDING_ENDPOINT = os.getenv("AZURE_OPENAI_EMBEDDING_ENDPOINT", "").rstrip("/")
AZURE_EMBEDDING_API_KEY = os.getenv("AZURE_OPENAI_EMBEDDING_API_KEY", "")
AZURE_EMBEDDING_DEPLOYMENT = os.getenv("AZURE_OPENAI_EMBEDDING_DEPLOYMENT", "text-embedding-3-small")
AZURE_EMBEDDING_API_VERSION = os.getenv("AZURE_OPENAI_EMBEDDING_API_VERSION", "2023-05-15")

QDRANT_URL = os.getenv("QDRANT_URL", "http://localhost:6333")
QDRANT_API_KEY = os.getenv("QDRANT_API_KEY") or None
COLLECTION_NAME = "kb_ggs_support"
AGENT_TYPE = "ggs_support"
DOC_FAMILY = "ecanter"
VECTOR_DIM = 1536

# Deployment budget measured from x-ratelimit-* headers: 250 RPM / 250k TPM.
# Held slightly under so a retry storm has headroom.
MAX_CONCURRENCY = 32
RPM_BUDGET = 220
TPM_BUDGET = 225_000
EST_TOKENS_PER_PAGE = 3_400

# Chunking. Extracted pages are already one-fact-per-line prose, so chunks are
# assembled from whole lines — a maintenance interval sentence is never cut in half.
CHUNK_SIZE = 1_200
CHUNK_OVERLAP_LINES = 2

# The legacy North America eCanter 1.0 sheet already in kb_ggs_support. Its
# 60–80 mile range figure predates every manual in this folder.
LEGACY_FILENAME = "ecanter_specsheet1.pdf"
LEGACY_PREFIX = (
    "[LEGACY DOCUMENT — eCanter 1.0, 2018, North America market. These figures are "
    "superseded by the current-generation eCanter manuals. State that this is the "
    "first-generation specification if you quote it.]\n\n"
)

# Human titles so a retrieved chunk names its own source in the agent's context.
DOC_TITLES = {
    "ASCENT HelpDocument.pdf": "FUSO ASCENT Portal Help Module (parts & service software)",
    "DTOM.pdf": "Daimler Truck Diagnostics Software (DTOM) Operation Manual",
    "HEV & EV.pdf": "eCanter HEV & EV Systems Guide",
    "Maintenance Manual For Canter.pdf": "Canter Maintenance Manual",
    "Owners Manual.pdf": "eCanter Owner's Manual",
    "Parts Catalogue SBOM.pdf": "Canter Parts Catalogue (SBOM)",
    "Service Bulletins.pdf": "FUSO Service Bulletins",
    "Service Manual.pdf": "Canter Service Manual",
    "Warranty Manual.pdf": "Canter Warranty Manual",
    "FEAVK_Damage_Code_Data.xlsx": "FEAVK Damage Code Reference Data",
}

# ─── Prompts ─────────────────────────────────────────────────────────────

LEGEND_SYSTEM = """You are reading pages from a Mitsubishi FUSO / Daimler Truck technical manual that define the symbols and abbreviations used throughout the document.

Extract ONLY the symbol/abbreviation definitions you can actually see. For each one give the symbol exactly as printed and its full meaning, one per line, as "SYMBOL = meaning".

Pay special attention to:
- Single letters used in maintenance schedule tables (I, R, C, L, T, A, etc.)
- Marks such as a star, a circle, a filled dot, a cross, a triangle, a square
- Whether a symbol denotes its OWN table column (for example a star column meaning "1st maintenance at 2,000 km") rather than an action taken at a distance interval. If a symbol heads its own column, say so explicitly on that line.

CRITICAL — THE DOT. These manuals typeset their legend as a bulleted list, so a
legend entry that DEFINES the dot looks identical to a bullet point. A line such as
"• : No inspection is necessary." is a DEFINITION of the dot symbol, not a list item.
If any line defines a dot, a filled circle or a bullet as meaning "no inspection",
"no action" or similar, you MUST emit it as: • = <meaning>. Never omit it.

If a page defines no symbols, output nothing for it. Output only the definition lines, no commentary."""

PAGE_SYSTEM = """You extract the FULL content of a page from a Mitsubishi FUSO / Daimler Truck technical document, rewriting it so every fact stands on its own when read in isolation by a support agent who cannot see the page.

Begin your output with one line: "SECTION: <the page's heading or topic>".

TABLES — THE MOST IMPORTANT RULE
Never output a raw grid, and never guess at alignment. For each table, work in this exact order:

 1. Write a line "COLUMNS: " listing every column header left to right, exactly as printed,
    INCLUDING any unlabelled or symbol-only column. A column whose header is a bare symbol
    (for example ☆) is a REAL column and is never a distance or time interval.

 2. For EACH data row, first write a working line of the exact form:
        CELLS: <row subject> | <column 2 header>=<cell value> | <column 3 header>=<cell value> | ...
    listing one entry for EVERY column you named in step 1, in the same left-to-right order,
    with the cell value transcribed literally (a dot stays "•", a letter stays "L").
    Count them: the number of entries after the row subject MUST equal the number of columns
    after the first. If a row visibly has fewer marks than there are columns, that means a cell
    is blank or merged — say "merged" or "blank", never silently shift the remaining values left.
    THE MOST COMMON AND MOST DAMAGING ERROR IS SHIFTING A ROW ONE COLUMN LEFT BY SKIPPING
    A SYMBOL-HEADED FIRST COLUMN. Do not do it.

 3. On the next line write that row as ONE self-contained sentence that repeats the row's
    subject and names each column it reports, translating every symbol through the LEGEND.

Resolve every symbol using the LEGEND block supplied below — a symbol's meaning comes ONLY from that legend, never from your own assumptions. If a symbol is genuinely absent from the legend, transcribe it literally and say its meaning is not defined; never invent one. Restate units and both metric and imperial figures exactly as printed.

Worked example for a row reading "Brake fluid | • | • | R | • | R: 24":
CELLS: Brake fluid | ☆=• | 40 (24)=• | 80 (48)=R | 120 (72)=• | Months=R: 24
Brake fluid: not part of the 1st maintenance at 2,000 km; no inspection necessary at 40,000 km (24,000 miles); Replace at 80,000 km (48,000 miles); no inspection necessary at 120,000 km (72,000 miles); and Replace every 24 months.

OTHER RULES
- Extract every heading, body paragraph, callout, figure label, note, footnote and warning on the page.
- Preserve EXACT part numbers, diagnostic trouble codes, torque values, capacities, voltages and model names. Never round or paraphrase a number.
- Use the PAGE TEXT LAYER as the authority for exact spelling of codes and numbers; use the IMAGE as the authority for layout, table structure, merged cells, and which symbol sits in which cell. If they disagree about a symbol or a position, trust the image.
- Software UI screenshots: state the named UI elements and what the caption says they do, phrased as instructions a user could follow. Do not describe visual styling.
- If the page carries no meaningful content (blank, or only a page number), output exactly: [BLANK PAGE]
- Output only extracted content. No commentary, no markdown fences."""


# ─── Rate limiting ───────────────────────────────────────────────────────


class RateLimiter:
    """Sliding-window limiter over both requests/min and tokens/min."""

    def __init__(self, rpm: int, tpm: int):
        self.rpm, self.tpm = rpm, tpm
        self._reqs: deque[float] = deque()
        self._toks: deque[tuple[float, int]] = deque()
        self._lock = threading.Lock()

    def acquire(self, est_tokens: int):
        while True:
            with self._lock:
                now = time.monotonic()
                while self._reqs and now - self._reqs[0] > 60:
                    self._reqs.popleft()
                while self._toks and now - self._toks[0][0] > 60:
                    self._toks.popleft()
                tok_sum = sum(t for _, t in self._toks)
                if len(self._reqs) < self.rpm and tok_sum + est_tokens < self.tpm:
                    self._reqs.append(now)
                    self._toks.append((now, est_tokens))
                    return
            time.sleep(0.35)


_limiter = RateLimiter(RPM_BUDGET, TPM_BUDGET)
_stats = {"in": 0, "out": 0, "calls": 0, "retries": 0}
_stats_lock = threading.Lock()


def _post_json(url: str, payload: dict, headers: dict, timeout: int) -> dict:
    req = urllib.request.Request(url, data=json.dumps(payload).encode(), headers=headers)
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.loads(resp.read())


def call_vision(b64_image: str, user_text: str, system: str, max_tokens: int = 4000, retries: int = 5) -> str:
    """One vision call, rate-limited, with exponential backoff on 429/5xx."""
    payload = {
        "messages": [
            {"role": "system", "content": system},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": user_text},
                    {
                        "type": "image_url",
                        "image_url": {"url": f"data:image/png;base64,{b64_image}", "detail": "high"},
                    },
                ],
            },
        ],
        "max_tokens": max_tokens,
        "temperature": 0.0,
    }
    last_err = None
    for attempt in range(retries):
        _limiter.acquire(EST_TOKENS_PER_PAGE)
        try:
            data = _post_json(
                VISION_ENDPOINT,
                payload,
                {"api-key": AZURE_API_KEY, "Content-Type": "application/json"},
                timeout=240,
            )
            usage = data.get("usage", {})
            with _stats_lock:
                _stats["in"] += usage.get("prompt_tokens", 0)
                _stats["out"] += usage.get("completion_tokens", 0)
                _stats["calls"] += 1
            return data["choices"][0]["message"]["content"] or ""
        except urllib.error.HTTPError as e:
            last_err = e
            body = ""
            try:
                body = e.read().decode()[:200]
            except Exception:
                pass
            if e.code == 429:
                wait = float(e.headers.get("Retry-After", 0) or 2 ** (attempt + 1))
            elif 500 <= e.code < 600:
                wait = 2 ** (attempt + 1)
            else:
                raise RuntimeError(f"vision HTTP {e.code}: {body}") from e
            with _stats_lock:
                _stats["retries"] += 1
            time.sleep(min(wait, 60))
        except Exception as e:  # timeouts, connection resets
            last_err = e
            with _stats_lock:
                _stats["retries"] += 1
            time.sleep(min(2 ** (attempt + 1), 60))
    raise RuntimeError(f"vision failed after {retries} attempts: {last_err}")


# ─── PDF helpers ─────────────────────────────────────────────────────────


def page_png_b64(doc, index: int) -> str:
    """Render a page, stepping DPI down until it fits comfortably in a request."""
    img_bytes = b""
    for dpi in (170, 140, 110, 85, 70):
        img_bytes = doc[index].get_pixmap(dpi=dpi).tobytes("png")
        if len(img_bytes) < 3_200_000:
            break
    return base64.b64encode(img_bytes).decode()


LEGEND_HINT = re.compile(
    r"symbol|legend|abbreviat|marks?\s+(used|given)|inspect\s*[:：]|"
    r"lubricat|replace or change|meaning of|how to (read|use) this",
    re.I,
)

# "• : No inspection is necessary." — the dot's own definition, which vision
# routinely mistakes for a bullet point. Captured straight from the text layer.
DOT_DEF = re.compile(r"^[•·●]\s*[:：]\s*(No\s+\w[^\n]{0,80})$", re.I | re.M)


def find_legend_pages(doc, limit: int = 3) -> list[int]:
    """Pages most likely to carry the document's symbol table.

    Legends live in front matter, so the search is biased there, but a scan of
    the whole document catches manuals that define symbols beside the first
    schedule table instead.
    """
    scored: list[tuple[int, int]] = []
    for i in range(doc.page_count):
        text = doc[i].get_text()
        hits = len(LEGEND_HINT.findall(text))
        if not hits:
            continue
        score = hits * 10
        if i < 30:
            score += (30 - i)  # front matter is the usual home
        scored.append((score, i))
    scored.sort(key=lambda s: (-s[0], s[1]))
    return [i for _, i in scored[:limit]]


def discover_legend(doc) -> str:
    """Read the document's own symbol definitions so page prompts can resolve cells."""
    pages = find_legend_pages(doc)
    if not pages:
        return ""
    fragments: list[str] = []
    for idx in pages:
        try:
            text = doc[idx].get_text().strip()[:4000]
            out = call_vision(
                page_png_b64(doc, idx),
                f"PAGE TEXT LAYER:\n---\n{text}\n---\n\n"
                "List every symbol/abbreviation definition visible on this page.",
                LEGEND_SYSTEM,
                max_tokens=1200,
            ).strip()
            if out and "=" in out:
                fragments.append(out)
        except Exception as e:
            print(f"      ⚠️  legend probe p{idx+1} failed: {e}", flush=True)

    if not fragments:
        return ""

    # De-duplicate definition lines across the probed pages, keeping first wins.
    # Note the strip set deliberately excludes "•": a leading dot here is usually
    # the symbol being defined, not list markup.
    seen: set[str] = set()
    lines: list[str] = []
    for frag in fragments:
        for line in frag.splitlines():
            line = line.strip(" -\t")
            if "=" not in line or len(line) > 300:
                continue
            key = line.split("=", 1)[0].strip().lower()
            if key and key not in seen:
                seen.add(key)
                lines.append(line)

    # Safety net for the dot. Vision reliably reads a legend line like
    # "• : No inspection is necessary." as a bullet and drops it, which leaves
    # every "no action" cell in the schedule unresolved. Recover it from the
    # text layer rather than assuming a meaning.
    if not any(l.lstrip().startswith(("•", "·", "●")) for l in lines):
        for i in range(min(doc.page_count, 60)):
            m = DOT_DEF.search(doc[i].get_text())
            if m:
                lines.append(f"• = {m.group(1).strip().rstrip('.')}")
                break
    return "\n".join(lines)


# ─── Phase 1: extraction ─────────────────────────────────────────────────


def load_checkpoint() -> dict[tuple[str, int], dict]:
    done: dict[tuple[str, int], dict] = {}
    if CHECKPOINT.exists():
        with CHECKPOINT.open() as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                    done[(rec["filename"], rec["page"])] = rec
                except json.JSONDecodeError:
                    continue
    return done


_write_lock = threading.Lock()


def append_checkpoint(rec: dict):
    with _write_lock:
        with CHECKPOINT.open("a") as fh:
            fh.write(json.dumps(rec, ensure_ascii=False) + "\n")
            fh.flush()


def extract_document(pdf_path: Path, done: dict, legend_cache: dict) -> int:
    filename = pdf_path.name
    doc = pymupdf.open(str(pdf_path))
    total = doc.page_count
    pending = [i for i in range(total) if (filename, i + 1) not in done]

    print(f"\n{'─' * 62}")
    print(f"📄 {filename}  ({total} pages, {len(pending)} to extract)")
    if not pending:
        print("   ✓ already complete in checkpoint")
        doc.close()
        return 0

    if filename in legend_cache:
        legend = legend_cache[filename]
    else:
        print("   🔎 discovering document legend...", end=" ", flush=True)
        legend = discover_legend(doc)
        legend_cache[filename] = legend
        print(f"{len(legend.splitlines())} symbol definitions" if legend else "none found")
    if legend:
        for line in legend.splitlines()[:12]:
            print(f"      · {line}")

    legend_block = (
        f"LEGEND / SYMBOLS DEFINED BY THIS DOCUMENT (authoritative — resolve every table cell with these):\n{legend}"
        if legend
        else "LEGEND: this document supplied no symbol table. If a cell holds a bare symbol you cannot "
        "resolve from the page itself, transcribe the symbol literally and say its meaning is not defined "
        "on this page. Do NOT invent a meaning."
    )
    title = DOC_TITLES.get(filename, filename)

    # A PyMuPDF Document is not thread-safe, so each worker thread opens its own
    # handle. Rendering then happens inside the worker and overlaps with every
    # other worker's network wait, instead of stalling the whole batch up front.
    local = threading.local()

    def handle():
        if not hasattr(local, "doc"):
            local.doc = pymupdf.open(str(pdf_path))
        return local.doc

    def work(index: int) -> dict:
        own = handle()
        b64 = page_png_b64(own, index)
        text_layer = own[index].get_text().strip()
        prompt = (
            f"DOCUMENT: {title}\nPAGE: {index + 1} of {total}\n\n"
            f"{legend_block}\n\n"
            f"PAGE TEXT LAYER (authority for exact strings):\n---\n{text_layer[:6000]}\n---\n\n"
            "Now extract this page per your rules."
        )
        out = call_vision(b64, prompt, PAGE_SYSTEM).strip()
        return {"filename": filename, "page": index + 1, "title": title, "text": out}

    written = 0
    completed = 0
    t_doc = time.monotonic()
    with ThreadPoolExecutor(max_workers=MAX_CONCURRENCY) as pool:
        futures = {pool.submit(work, i): i for i in pending}
        for fut in as_completed(futures):
            index = futures[fut]
            completed += 1
            try:
                rec = fut.result()
            except Exception as e:
                print(f"\n   ❌ p{index+1}: {e}", flush=True)
                continue
            if rec["text"] and rec["text"] != "[BLANK PAGE]":
                append_checkpoint(rec)
                written += 1
            else:
                append_checkpoint({**rec, "text": ""})
            if completed % 25 == 0 or completed == len(pending):
                rate = completed / max(0.01, time.monotonic() - t_doc) * 60
                eta = (len(pending) - completed) / max(0.1, rate)
                print(
                    f"   ⏳ {completed}/{len(pending)} ({completed/len(pending)*100:.0f}%)  "
                    f"{rate:.0f} pages/min  eta {eta:.0f}m  retries={_stats['retries']}",
                    flush=True,
                )

    doc.close()
    return written


def extract_xlsx(path: Path, done: dict) -> int:
    """Damage-code workbook → one self-describing sentence per row."""
    filename = path.name
    if any(k[0] == filename for k in done):
        print(f"\n📊 {filename}  ✓ already in checkpoint")
        return 0
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    title = DOC_TITLES.get(filename, filename)
    written = 0
    print(f"\n📊 {filename}  ({len(wb.sheetnames)} sheet(s))")
    for si, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            header = [str(c).strip() if c is not None else "" for c in next(rows)]
        except StopIteration:
            continue
        lines: list[str] = []
        for row in rows:
            parts = [
                f"{header[i]}: {str(v).strip()}"
                for i, v in enumerate(row)
                if i < len(header) and v is not None and str(v).strip() and header[i]
            ]
            if parts:
                lines.append("; ".join(parts) + ".")
        if not lines:
            continue
        append_checkpoint(
            {
                "filename": filename,
                "page": si + 1,
                "title": title,
                "text": f"SECTION: {sheet_name} — damage code reference\n" + "\n".join(lines),
            }
        )
        written += 1
        print(f"   ✓ sheet '{sheet_name}': {len(lines)} rows")
    wb.close()
    return written


# ─── Phase 2: chunk, embed, upsert ───────────────────────────────────────


def chunk_page(text: str, title: str, page: int) -> list[str]:
    """Assemble chunks from whole lines and stamp each with its own context.

    Every chunk is prefixed with the document title and the page's SECTION line,
    so a chunk retrieved on its own still says which manual and which procedure
    it came from — the agent reads that context aloud rather than guessing.
    """
    lines = [l.rstrip() for l in text.splitlines()]
    section = ""
    if lines and lines[0].upper().startswith("SECTION:"):
        section = lines[0][8:].strip()
        lines = lines[1:]
    # "CELLS:" lines are the model's working — forcing it to account for every
    # column is what keeps a row from shifting left, but the agent should read
    # the prose sentence that follows, not the scaffolding.
    lines = [l for l in lines if l.strip() and not l.lstrip().upper().startswith("CELLS:")]
    if not lines:
        return []

    header = f"[{title} — page {page}"
    header += f" — {section}]" if section else "]"

    chunks: list[str] = []
    current: list[str] = []
    size = 0
    for line in lines:
        # A single line longer than the budget becomes its own chunk rather than
        # being split mid-sentence.
        if size + len(line) > CHUNK_SIZE and current:
            chunks.append(header + "\n" + "\n".join(current))
            current = current[-CHUNK_OVERLAP_LINES:] if len(current) > CHUNK_OVERLAP_LINES else []
            size = sum(len(l) for l in current)
        current.append(line)
        size += len(line) + 1
    if current:
        chunks.append(header + "\n" + "\n".join(current))
    return chunks


def embed_batch(texts: list[str], retries: int = 5) -> list[list[float]]:
    url = (
        f"{AZURE_EMBEDDING_ENDPOINT}/openai/deployments/{AZURE_EMBEDDING_DEPLOYMENT}"
        f"/embeddings?api-version={AZURE_EMBEDDING_API_VERSION}"
    )
    for attempt in range(retries):
        try:
            data = _post_json(
                url,
                {"input": [t[:8000] for t in texts]},
                {"api-key": AZURE_EMBEDDING_API_KEY, "Content-Type": "application/json"},
                timeout=120,
            )
            return [d["embedding"] for d in sorted(data["data"], key=lambda x: x["index"])]
        except urllib.error.HTTPError as e:
            if e.code == 429 or 500 <= e.code < 600:
                time.sleep(min(2 ** (attempt + 1), 45))
                continue
            raise
        except Exception:
            time.sleep(min(2 ** (attempt + 1), 45))
    raise RuntimeError("embedding failed after retries")


def qdrant_client():
    from qdrant_client import QdrantClient

    return QdrantClient(url=QDRANT_URL, api_key=QDRANT_API_KEY, timeout=60.0)


def with_retry(what: str, fn, retries: int = 12):
    """Retry a Qdrant call through transient network faults.

    Qdrant Cloud sits behind a DNS name with several A records; a momentary
    resolver failure or a dropped connection surfaces as an exception that is
    fatal to a long ingest but succeeds on the very next attempt. Losing an
    hour of work to one blip is not acceptable, so every write goes through here.
    """
    last = None
    for attempt in range(retries):
        try:
            return fn()
        except Exception as e:
            last = e
            wait = min(2 ** attempt, 45)
            print(f"      ⚠️  {what} failed ({type(e).__name__}); retry {attempt+1}/{retries} in {wait}s", flush=True)
            time.sleep(wait)
    raise RuntimeError(f"{what} failed after {retries} attempts: {last}")


def ensure_collection(client):
    from qdrant_client.models import Distance, PayloadSchemaType, VectorParams

    existing = [c.name for c in client.get_collections().collections]
    if COLLECTION_NAME not in existing:
        client.create_collection(
            collection_name=COLLECTION_NAME,
            vectors_config=VectorParams(size=VECTOR_DIM, distance=Distance.COSINE),
        )
        print(f"  ✅ created collection '{COLLECTION_NAME}'")
    else:
        info = client.get_collection(COLLECTION_NAME)
        print(f"  ✓ collection '{COLLECTION_NAME}' exists ({info.points_count} points)")

    # Filtering by payload needs a keyword index on the server — without it
    # Qdrant rejects the per-filename delete that makes re-runs idempotent, and
    # the legacy-tagging scroll. Creating an existing index is a no-op.
    indexes = {
        "filename": PayloadSchemaType.KEYWORD,
        "doc_family": PayloadSchemaType.KEYWORD,
        "agent_type": PayloadSchemaType.KEYWORD,
        # `page` is filtered alongside filename to pull a whole page back
        # together when an answer spans several chunks.
        "page": PayloadSchemaType.INTEGER,
    }
    for field, schema in indexes.items():
        try:
            client.create_payload_index(
                collection_name=COLLECTION_NAME,
                field_name=field,
                field_schema=schema,
                wait=True,
            )
            print(f"  ✓ payload index on '{field}'")
        except Exception as e:
            print(f"  ⚠️  payload index on '{field}': {e}")


def delete_by_filename(client, filename: str):
    """Idempotency: a re-run replaces a document's vectors instead of duplicating them."""
    from qdrant_client.models import FieldCondition, Filter, MatchValue

    with_retry(
        f"delete '{filename}'",
        lambda: client.delete(
            collection_name=COLLECTION_NAME,
            points_selector=Filter(
                must=[FieldCondition(key="filename", match=MatchValue(value=filename))]
            ),
            wait=True,
        ),
    )


def embed_and_upsert(records: list[dict], client) -> int:
    from qdrant_client.models import PointStruct

    by_file: dict[str, list[dict]] = {}
    for rec in records:
        by_file.setdefault(rec["filename"], []).append(rec)

    grand_total = 0
    for filename, recs in by_file.items():
        recs.sort(key=lambda r: r["page"])
        # chunk_index is the chunk's ordinal WITHIN its page. It lets a reader
        # reassemble a page in printed order from whichever chunk matched, which
        # is how a procedure or a schedule table is served whole instead of as
        # the one fragment that happened to score best.
        chunks: list[tuple[str, int, int]] = []
        for rec in recs:
            for idx, chunk in enumerate(chunk_page(rec["text"], rec["title"], rec["page"])):
                chunks.append((chunk, rec["page"], idx))
        if not chunks:
            continue

        print(f"\n  📦 {filename}: {len(recs)} pages → {len(chunks)} chunks")
        delete_by_filename(client, filename)

        points = []
        batch = 16
        t0 = time.monotonic()
        for i in range(0, len(chunks), batch):
            group = chunks[i : i + batch]
            vectors = embed_batch([c for c, _, _ in group])
            for (chunk_text, page, chunk_index), vec in zip(group, vectors):
                points.append(
                    PointStruct(
                        id=str(uuid.uuid4()),
                        vector=vec,
                        payload={
                            "text": chunk_text,
                            "filename": filename,
                            "title": DOC_TITLES.get(filename, filename),
                            "page": page,
                            "chunk_index": chunk_index,
                            "agent_type": AGENT_TYPE,
                            "category": "ecanter_technical",
                            "doc_family": DOC_FAMILY,
                        },
                    )
                )
            if len(points) >= 200:
                batch_points = points
                with_retry(
                    f"upsert {len(batch_points)} pts",
                    lambda bp=batch_points: client.upsert(
                        collection_name=COLLECTION_NAME, points=bp, wait=False
                    ),
                )
                points = []
            if (i // batch) % 20 == 0 and i:
                done_n = min(i + batch, len(chunks))
                rate = done_n / max(0.01, time.monotonic() - t0) * 60
                print(f"     … {done_n}/{len(chunks)} embedded ({rate:.0f}/min)", flush=True)
        if points:
            tail_points = points
            with_retry(
                f"upsert {len(tail_points)} pts (final)",
                lambda: client.upsert(
                    collection_name=COLLECTION_NAME, points=tail_points, wait=True
                ),
            )
        grand_total += len(chunks)
        print(f"     ✅ {len(chunks)} vectors upserted")
    return grand_total


# ─── Legacy specsheet tagging ────────────────────────────────────────────


def tag_legacy_specsheet(client) -> int:
    """Prefix the 2018 eCanter 1.0 vectors so the agent dates them out loud.

    The vectors are kept and re-embedded with the prefix in place, so the notice
    is part of what similarity search returns rather than a payload field the
    agent would never see.
    """
    from qdrant_client.models import FieldCondition, Filter, MatchValue, PointStruct

    flt = Filter(must=[FieldCondition(key="filename", match=MatchValue(value=LEGACY_FILENAME))])
    points, _ = client.scroll(
        collection_name=COLLECTION_NAME, scroll_filter=flt, limit=500, with_payload=True, with_vectors=False
    )
    if not points:
        print(f"  ℹ️  no '{LEGACY_FILENAME}' vectors found — nothing to tag")
        return 0

    todo = [p for p in points if not (p.payload or {}).get("text", "").startswith("[LEGACY DOCUMENT")]
    if not todo:
        print(f"  ✓ '{LEGACY_FILENAME}' already tagged as legacy ({len(points)} vectors)")
        return 0

    print(f"  🏷️  tagging {len(todo)} '{LEGACY_FILENAME}' vectors as legacy")
    rebuilt = []
    for i in range(0, len(todo), 16):
        group = todo[i : i + 16]
        texts = [LEGACY_PREFIX + (p.payload or {}).get("text", "") for p in group]
        vectors = embed_batch(texts)
        for p, text, vec in zip(group, texts, vectors):
            payload = dict(p.payload or {})
            payload["text"] = text
            payload["doc_family"] = "ggs_core_legacy"
            payload["superseded_by"] = "Owners Manual.pdf, Service Manual.pdf, HEV & EV.pdf"
            rebuilt.append(PointStruct(id=p.id, vector=vec, payload=payload))
    with_retry(
        "upsert legacy vectors",
        lambda: client.upsert(collection_name=COLLECTION_NAME, points=rebuilt, wait=True),
    )
    print(f"  ✅ {len(rebuilt)} legacy vectors updated")
    return len(rebuilt)


# ─── Main ────────────────────────────────────────────────────────────────


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--extract-only", action="store_true")
    ap.add_argument("--embed-only", action="store_true")
    ap.add_argument("--tag-legacy", action="store_true", help="only re-tag the old eCanter specsheet")
    ap.add_argument("--files", default="", help="comma-separated filenames to limit the run to")
    args = ap.parse_args()

    if not AZURE_API_KEY:
        sys.exit("AZURE_OPENAI_CHAT_API_KEY is not set in bridge/.env")

    print("=" * 62)
    print("  E-Canter → kb_ggs_support ingestion")
    print(f"  vision={VISION_DEPLOYMENT}  embed={AZURE_EMBEDDING_DEPLOYMENT}  concurrency={MAX_CONCURRENCY}")
    print("=" * 62)

    if args.tag_legacy:
        client = qdrant_client()
        ensure_collection(client)
        tag_legacy_specsheet(client)
        return

    only = {f.strip() for f in args.files.split(",") if f.strip()}
    sources = sorted(
        p for p in SOURCE_DIR.iterdir() if p.suffix.lower() in (".pdf", ".xlsx") and not p.name.startswith("~$")
    )
    if only:
        sources = [p for p in sources if p.name in only]
    if not sources:
        sys.exit(f"no source documents found in {SOURCE_DIR}")

    # ── Phase 1 ──
    if not args.embed_only:
        done = load_checkpoint()
        print(f"\n📁 {len(sources)} source documents · checkpoint holds {len(done)} pages")
        legend_cache: dict[str, str] = {}
        t0 = time.monotonic()
        for path in sources:
            if path.suffix.lower() == ".xlsx":
                extract_xlsx(path, done)
            else:
                extract_document(path, done, legend_cache)
        mins = (time.monotonic() - t0) / 60
        cost = _stats["in"] / 1e6 * 0.40 + _stats["out"] / 1e6 * 1.60
        print(f"\n{'=' * 62}")
        print(f"  extraction done in {mins:.1f} min · {_stats['calls']} vision calls")
        print(f"  tokens: {_stats['in']:,} in + {_stats['out']:,} out  (~${cost:.2f})")
        print(f"  retries: {_stats['retries']}")

    if args.extract_only:
        print("\n--extract-only set; stopping before embedding.")
        return

    # ── Phase 2 ──
    records = [r for r in load_checkpoint().values() if r.get("text", "").strip()]
    if only:
        records = [r for r in records if r["filename"] in only]
    if not records:
        sys.exit("checkpoint is empty — run extraction first")

    print(f"\n🔗 embedding {len(records)} extracted pages into '{COLLECTION_NAME}'")
    client = qdrant_client()
    ensure_collection(client)
    total = embed_and_upsert(records, client)

    if not only:
        print("\n🏷️  legacy specsheet")
        tag_legacy_specsheet(client)

    # Reference dump of everything the vision pass read.
    with EXTRACT_DUMP.open("w") as fh:
        for rec in sorted(records, key=lambda r: (r["filename"], r["page"])):
            fh.write(f"{'=' * 62}\nFILE: {rec['filename']}  PAGE: {rec['page']}\n{'=' * 62}\n")
            fh.write(rec["text"] + "\n\n")
    print(f"\n💾 extracted text dumped to {EXTRACT_DUMP}")

    info = client.get_collection(COLLECTION_NAME)
    print(f"\n{'=' * 62}")
    print(f"  ✅ done — {total} E-Canter vectors added")
    print(f"  '{COLLECTION_NAME}' now holds {info.points_count} points")
    print("=" * 62)


if __name__ == "__main__":
    main()
